#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Copyright Huawei Technologies Co., Ltd. 2010-2020. All rights reserved.
"""
5.设置行高和列宽
6.合并单元格
"""
from openpyxl import load_workbook
from example import test_dir

if __name__ == "__main__":
    # 1.加载创建的工作簿
    wb = load_workbook(f"{test_dir}\\7月下旬入库表.xlsx")
    ws = wb.active  # 等于 wb["7月下旬入库表"]，只要是默认第一张表都可以用wb.active

    # 写入单元格
    ws["F1"] = "默认"
    ws["G2"] = "设置行高"
    ws["H3"] = "设置列宽"

    # 设置第二行行高
    ws.row_dimensions[2].height = 40
    # 设置H列列狂
    ws.column_dimensions["C"].width = 30

    # 合并单元格
    ws.merge_cells("A2:D4")
    # 等同于 ws.merge_cells(star_row=2,start_colum=1,end_row=4,end_column=4)

    ws["A2"] = "合并单元格"  # 在合并区域的左上角A2写入

    # 取消合并
    ws.unmerge_cells("A2:D4")
    # 等同于 ws.unmerge_cells(star_row=2,start_colum=1,end_row=4,end_column=4)
