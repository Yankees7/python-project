#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Copyright Huawei Technologies Co., Ltd. 2010-2020. All rights reserved.
"""
写入数据
"""
from openpyxl import load_workbook
from example import test_dir
from openpyxl.drawing.image import Image

if __name__ == "__main__":
    # 1.加载创建的工作簿
    wb = load_workbook(f"{test_dir}\\7月下旬入库表.xlsx")
    ws = wb.active  # 等于 wb["7月下旬入库表"]，只要是默认第一张表都可以用wb.active

    """写入数据"""
    # 在单元格写入数据
    ws["J1"] = 42  # 在J1单元格写入
    ws.cell(row=2, column=11, value=42)  # 在K2单元格写入
    ws.cell(3, 11).value = 42  # 在K1单元格写入

    # 新增一行数据
    ws.append([1, 2, 3, 4])

    """写入公式"""
    ws["B2"] = "=SUM(A2:A4)"
    ws.cell(row=2, column=2, value="SUM(A2:B4)")
    ws.cell(2, 2).value = "=SUM(A2:A4)"

    """插入图片"""
    img = Image(f"{test_dir}\\default.jpg")  # image:要插入的图片
    ws.add_image(img, "M2")  # 在B1单元格插入图片
    # 设置行高列宽以更好显示图片
    ws.row_dimensions[2].height = 40
    ws.column_dimensions["M"].width = 10

    wb.save(f"{test_dir}\\7月下旬入库表-副本.xlsx")
