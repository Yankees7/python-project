#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Copyright Huawei Technologies Co., Ltd. 2010-2020. All rights reserved.
"""
4.访问单元格
当一个工作表被创建时，启动不包含单元格，只有单元格被获取时才被创建。
这种方式我们不会创建我们不会使用的单元格，从而减少了内存消耗
"""
from openpyxl import load_workbook
from example import test_dir

if __name__ == "__main__":
    # 1.加载创建的工作簿
    wb = load_workbook(f"{test_dir}\\7月下旬入库表.xlsx")
    ws = wb.active  # 等于 wb["7月下旬入库表"]，只要是默认第一张表都可以用wb.active
    """
    1.单个单元格访问
    """
    # 方法1.
    cell_A2 = ws["A2"]
    # 方法2
    cell_B2 = ws.cell(row=2, column=2)
    # 返回单元格对象
    print(cell_A2, cell_B2)
    """
    2.多个单元格访问
    结果都可以用tuple(),list(),循环进行处理
    """
    # 通过切片
    cell_area = ws["A1":"B4"]
    cell_exact = ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2)
    print(cell_area)  # 每一行为元素（每一行的每个单元格对象组成的元组）组成的元组
    for rows in cell_area:
        for cells in rows:
            print(cells.value)
    # 通过行列
    col_A = ws["A"]  # A列
    col_area = ws["A":"B"]  # A，B列
    row_2 = ws[2]  # 第二行
    row_area = ws[2:5]  # 2-5行

    # 迭代所有行
    all_by_row = ws.rows

    # 叠戴所有列
    all_by_col = ws.columns
