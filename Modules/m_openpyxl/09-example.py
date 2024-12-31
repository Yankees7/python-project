from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, NamedStyle
from openpyxl import Workbook
from copy import copy

wb = Workbook()
ws = wb.active

top_left_cell = ws["B2"]  # 选定单元格B2
top_left_cell.value = "设置样式"  # 写入内容

"""
要想改变合并单元格的样式，也只需要改变合并区域左上角的单元格样式即可。

注意：如果先合并单元格，再设置样式，合并单元格的边框设置不完整。而如果是先设置样式再合并单元格，则没有问题。
"""
# 设置格式
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
top_left_cell.border = Border(top=double, left=thin, right=thin, bottom=double)
top_left_cell.fill = PatternFill("solid", fgColor="DDDDDD")
top_left_cell.font = Font(b=True, color="FF0000")
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

ws.merge_cells("B2:F4")  # 合并单元格

wb.save("实例.xlsx")

# 复制样式

ft1 = Font(name="Arial", size=14)
ft2 = copy(ft1)
ft2.name = "Tahoma"

print(ft1.name)  # Arial
print(ft2.name)  # Tahoma
print(ft2.size)  # 14.0

"""
应用样式
"""
# 我们上面的样式都是直接应用于单元格
ws["A1"].font = Font(size=12)

# 样式也可以应用于整行和整列，整行整列应用后对已有的单元格数据并不会发生改变，只有在手动打开表格新增内容时才会改变
ws["A1"] = 12
ws["A2"] = 12

col = ws.column_dimensions["A"]
col.font = Font(bold=True)

row = ws.row_dimensions[1]
row.font = Font(underline="single")

"""
创建命名样式
"""

highlight = NamedStyle(name="highlight")
highlight.font = Font(bold=True, size=20)
bk = Side(style="thick", color="000000")
highlight.border = Border(left=bk, top=bk, right=bk, bottom=bk)

# 使用
wb.add_named_style(highlight)
ws["B2"] = 2
ws["D5"] = 5

ws["B2"].style = highlight


#
from openpyxl.utils import get_column_letter, column_index_from_string


# 根据列的数字返回字母
print(get_column_letter(2))  # B

# 根据字母返回列的数字
print(column_index_from_string("D"))  # 4


"""颜色设置
字体/填充/边框的颜色可以通过三种方式来设置： indexed, aRGB or theme
一般建议用aRGB ，因为其他两种受excel限制。

"""
# aRGB颜色
font = Font(color="00FF9900")

# 索引颜色、主题颜色
from openpyxl.styles.colors import Color

font = Font(color=Color(indexed=32))
font = Font(color=Color(theme=6, tint=0.5))
