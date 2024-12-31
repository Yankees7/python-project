#!/usr/bin/env python
# -*- coding: UTF-8 -*-
# Copyright yangjing Technologies Co., Ltd. 2010-2020. All rights reserved.
'''
一、对工作簿中对应工作表的相关操作
'''

from openpyxl import load_workbook
from example import test_dir

if __name__ == "__main__":
    # 1.加载创建的工作簿
    wb = load_workbook(f"{test_dir}/新建工作簿.xlsx")
    # 2.读取工作表sheet1
    sh1 = wb['sheet1']
    # 打印表名
    print(sh1.title)
    # 3.修改sheet1重命名为newsheet1
    sh1.title = 'newsheet1'

    # 4.复制工作表
    newsh = wb.copy_worksheet(sh1)
    # 指定新工作表名字
    newsh.title = 'newhsheet'
    print(newsh.title)

    # 删除工作表
    wb.remove(sh1)
    '''查看工作簿里所有表名字'''
    print(wb.sheetnames)

    # 创建工作表sheet1
    wb.create_sheet('sheet1', 0)

    # 最后保存
    wb.save(f"{test_dir}/新建工作簿.xlsx")
