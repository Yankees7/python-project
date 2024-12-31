from openpyxl import Workbook
from openpyxl.styles import Protection

"""1、保护工作薄

防止其他用户查看隐藏的工作表，添加、移动或隐藏工作表以及重命名工作表，可以使用密码保护 Excel 工作簿的结构。

"""
wb = Workbook()
ws = wb.active

# 启动工作表保护,不指定密码
# ws.protection.sheet = True
# ws.protection.enable()

# 不启动工作表保护
# ws.protection.sheet = False
# ws.protection.disable()


ws.protection.sheet = True  # 启动工作表保护
ws.protection.password = "123"  # 设置保护密码

ws["A1"] = "已保护"  # A1单元格写入
ws["C3"] = "=1+2+3"  # C3单元格写入

# 设置保护格式
protection = Protection(locked=True, hidden=True)
ws["C3"].protection = protection  # 注意：只有保护工作表后，锁定单元格或隐藏公式才有效。
"""
locked    # 是否设置锁定，默认True。其他项：False
hidden   # 是否设置隐藏，默认False。其他项：True
"""
wb.save("实例.xlsx")
