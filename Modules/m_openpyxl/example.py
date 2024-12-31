#!/usr/bin/evn python
"""
初学openpyxl模块,对excel处理之创建工作簿
"""

from openpyxl import Workbook
import os
import pathlib

homedir = str(pathlib.Path.home())  # 获取家目录,跨系统
test_dir = os.path.join(homedir, "Desktop", "test")
if not os.path.exists(test_dir):
    os.makedirs(test_dir)

if __name__ == "__main__":
    # 创建工作簿对象wb
    wb = Workbook()

    # 激活 worksheet；值默认为 0。除非你修改了这个值，不然这个方法会一直获取第一个工作表。
    ws = wb.active

    # 创建工作表.工作薄在创建时会自动生成一个名字，以(Sheet, Sheet1, Sheet2, …)来进行命名
    ws1 = wb.create_sheet("Mysheet")  # insert at the end（default）
    ws2 = wb.create_sheet("Mysheet", 0)  # insert at first position
    ws3 = wb.create_sheet("Mysheet", -1)  # insert at the penultimate position
    ws4 = wb.create_sheet("sheet1", 0)  # insert at the penultimate position

    # 修改工作表名
    ws.title = "New title"

    # 修改工作表选项卡背景颜色;也可以使用 RRGGBB 颜色来改变 Worksheet.sheet_properties.tabColor 属性
    ws.sheet_properties.tabColor = "1072BA"

    # 查看工作簿中所有工作表名称
    print(wb.sheetnames)
    # 或者通过遍历实现查看工作簿中所有工作表名称
    for sheet in wb:
        print(sheet.title)

    # 复制ws工作表
    target = wb.copy_worksheet(ws)
    target.title = "Newworksheet"

    # 移动工作表
    wb.move_sheet(ws, +1)

    # （重要）最后保存工作簿
    wb.save(f"{test_dir}/新建工作簿.xlsx")
