from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import GradientFill
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment
from example import test_dir
import datetime


wb = Workbook()
ws = wb.active

"""
数字格式
"""
ws["A1"] = "文字"
print(ws["A1"].number_format)  # -->> General

ws["B2"] = 5
print(ws["B2"].number_format)

ws["A3"] = 0.05
ws["A3"].number_format = "0.00%"  # 自定义格式
print(ws["A3"].number_format)  # -->> 0.00%

ws["B1"] = datetime.datetime.now()
print(ws["B1"].number_format)  # -->> yyyy-mm-dd hh:mm:ss

ws["B2"] = datetime.datetime.now()
ws["B2"].number_format = "yyyy-mm-dd"  # 自定义格式
print(ws["B2"].number_format)  # -->> yyyy-mm-dd

"""
字体
参数说明：
name # 字体
size # 字号,默认11
bold # 是否加粗，默认False。加粗：True
italic # 是否斜体，默认False。斜体：True
verAlign # 上下标，默认None。正常：baseline，上标：superscript，下标：subscript
color # 字体颜色，默认黑色（FF000000）
strikethrough # 删除线，默认不设置。设置：True
underline # 下划线，默认不带下划线。单下划线：single，双下划线：double，会计用单下划线：singleAccounting，会计用双下划线：doubleAccounting
"""

ws["D1"] = "默认"  # D1单元格写入
ws["E2"] = "设置格式"  # E2单元格写入
ws["F3"] = "设置上标"  # F3单元格写入

# 设置E2单元格字体格式
ws["E2"].font = Font(
    name="Calibri",
    size=12,
    color="00FF9900",
    italic=True,
    underline="double",
    strikethrough=True,
)
# 设置F3单元格字体格式
ws["F3"].font = Font(vertAlign="superscript", bold=True)

"""
填充
"""
"""
纯色填充
fill_type=? #设置图案样式，如果不设置则不会显示颜色
solid：实心，lightHorizontal,darkTrellis,darkup,darkGray,darkVertical,lightDown,lightTrellis,lightUP,darkDown,darkHorizontal,mediumGray,lightVertical,gray0625,gray125,lighGrid,lightGary
fgColor/start_color #前景色，即填充色
bgColor/end_color # 背景色，即图案颜色
"""
ws["H1"] = "默认"
ws["I2"] = "前景色"
ws["J3"] = "背景色"

# 前景色即填充色。也是我们一般设置的填充色
ws["I2"].fill = PatternFill(fill_type="solid", fgColor="00FF9900")

# 背景色即图案颜色
ws["J3"].fill = PatternFill(fill_type="solid", bgColor="00FF9900")

wb.save(f"{test_dir}\\实例.xlsx")

"""
渐变填充
type/fill_type # 渐变填充类型：linear，path
linear：
渐变在一组指定的Stops之间插入颜色，跨越一个区域长度。默认情况下渐变式从左到右的，但可以使用degree属性修改此方向。可以改为提供颜色列表，他们之间的距离将相等。
path:
渐变从区域的每个边缘应用线性渐变。属性：top、right、bottom、left。指定从各个边界填充给你的范围。比如：top="2",将填充单元格的前20%
"""
ws.merge_cells("B2:F4")

# 对合并单元格左上角单元格设置渐变填充
top_left_cell = ws["B2"]
top_left_cell.fill = GradientFill(
    type="linear", degree=0, stop=("FFFFFF", "99ccff", "000000")
)

"""
Border参数说明
left = Side(style,color) 左边框设置
right = Side(style,color 右边框设置
top = Side(style,color) 上边框设置
bottom = Side(style,color) 下边框设置

diagonalDown 是否显示左上右下对角线，显示：True
diagonaUp 是否显示左下右上对角线，显示：True
diagonal = Side(style,color) 对角边框设置，注意首要先设置显示对角线

Side参数说明
style/border_style 边框样式
有：thick，mediumDashDot,dashed,mediumDashDotDot,dashDot,slantDashDot,dotted,thin,hair,dashDotDot,mediumDashed,medium

color 边框颜色

"""
ws["A4"] = "默认"
ws["B5"] = "边框"
ws["C6"] = "对角线"

# 边框线格式设置
line_format = Side(style="medium", color="00FF9900")

# B5单元格设置上下左右边框
ws["B5"].border = Border(
    left=line_format, right=line_format, top=line_format, bottom=line_format
)

# C6单元格设置对角线
ws["C6"].border = Border(diagonalDown=True, diagonalUp=True, diagonal=line_format)


"""
Alignment
参数说明：
horizontal 水平对齐方式，默认general.还可设置：center,fill,left,distributed,right,centerContinus,justify
vertical 垂直对齐方式，默认靠下bottom。还可设置：center,top,justify,distributed
text_rotation/textRotation 文字旋转，默认0°，可设置：-90-90°
wrap_text/wrapText 设置自动换行，默认不设置（False).设置：True
shrinkToFit/shrink_to_fit 设置缩小字体填充，默认不设置，设置：True
indent 缩进，默认0
readingOrder 文字方向，默认0。0: 根据内容，1；总是从左往右，2：总是从右往左
"""

ws["A8"] = "默认"  # A1单元格写入
ws["B9"] = "居中并且换行"  # B2单元格写入
ws["C10"] = "靠上不换行"  # C3单元格写入

ws["B9"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws["C10"].alignment = Alignment(horizontal="center", vertical="top")


ws["C10"].alignment = Alignment(horizontal="center", vertical="top")

wb.save(f"{test_dir}\\实例.xlsx")
