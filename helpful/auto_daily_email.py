"""
SmartKit自动化每日邮件codecheck版本级代码检查情况
步骤：
1.去Libing下载版本级统计表格：CleanCode→版本质量代码→质量分析→子系统/服务→旁边打勾全选→点击导出统计数据（第一步暂时手动下载，后续考虑做web自动化下载）
2.对下载的版本级统计表格的数据进行整理
3.对整理后的数据美化
"""

import os
import stat
import re
import time
import logging as log

import win32com.client as win32

from spire.xls import Workbook

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter  # 通过数字序号得到列的字母序号
from openpyxl.styles import Font, Alignment, colors, Border, Side


log.basicConfig(level=log.INFO)


class Daily_Email:
    # 当前C版本；注意更换，估计一年更换一次
    product_version = "tenxuncloud 24.0.0"

    # 获取当前日期，例如：20240701
    date_time = time.strftime("%Y%m%d")
    y_m_d = time.strftime("%Y年%m月%d日")

    # 下载的版本级统计表格名称
    Libing_xlsx = f"{product_version}_{date_time}.xlsx"

    # 排除一些模块名称
    exclude_module_list = []

    # 所需的列名称
    colume_names = [
        "模块名称",
        "缺陷总数",
        "最小集",
        "要求类",
        "建议类",
        "代码规模",
        "平均函数代码行",
        "平均文件代码行",
        "总文件重复率",
        "总代码重复率",
        "不安全函数（未处理）",
        "冗余代码（未处理）",
        "超大头文件（未处理）",
        "超大源代码文件（未处理）",
        "超大函数（未处理）",
        "超大圈复杂度（未处理）",
        "平均圈复杂度",
        "超大目录（未处理）",
        "超大深度（未处理）",
        "抑制告警数",
    ]

    @classmethod
    def process_libing_xlsx(cls):
        """处理数据"""
        # 加载工作簿
        wb = load_workbook(cls.Libing_xlsx)
        # 创建工作表，名字叫：Mysheet,位置0
        if "Mysheet" in wb.sheetnames:  # 表格如果已经存在，那就删除，在重新创建，清空历史数据
            del wb["Mysheet"]
        ws1 = wb.create_sheet("Mysheet", 0)
        # 读取工作表，名字叫：告警数据
        ws2 = wb["告警数据"]
        # 打印一下告警数据的行数和列数
        log.info("告警数据表的列数：{}".format(ws2.max_column))
        log.info("告警数据表的行数：{}".format(ws2.max_row))

        # 筛选需展示模块的行的序数row_num_list
        row_num_list = []  # 可用行数列表
        for row_num in range(1, ws2.max_row + 1):  # 为什么加1，因为range()包左不包右
            # 模块的坐标
            module_coord = f"D{str(row_num)}"
            if ws2[module_coord].value in cls.exclude_module_list:
                continue
            row_num_list.append(row_num)
        log.info(f"可用行的序数列表：{row_num_list}")

        # 筛选需要的列的序数column_num_list
        column_num_list = []
        for col_num in range(4, ws2.max_column + 1):  # 从第4开始，其实随便无所谓，只筛选需要的列
            for checkname in cls.colume_names:
                if checkname == ws2.cell(row=1, column=col_num).value:
                    column_num_list.append(col_num)
        log.info(f"可用列的序数列表：{column_num_list}")

        # 将原表格所需要的内容提取出来
        data_new = []  # 一行数据
        for r in row_num_list:
            for c in column_num_list:
                data_new.append(ws2.cell(row=r, column=c).value)  # 列表里面最佳值
            ws1.append(data_new)  # 向MySheet表写入
            data_new = []  # 写下一行之前，清空上一行

        # 插入索引
        ws1.insert_cols(1, 1)  # 第一列前插入1列索引列
        ws1["A1"].value = "序号"
        for seq in range(2, ws1.max_row + 1):  # 从第二行开始填充序号
            ws1[f"A{seq}"].value = seq - 1

        # 写入codecheck的数据
        ws1.insert_cols(3, 1)  # 第三列前插入1列
        ws1["C1"].value = "codecheck检查"

        # 保存工作簿
        wb.save(cls.Libing_xlsx)

    @classmethod
    def beautiful_libing_xlsx(cls):
        """进行美化"""
        # 字体
        warning_value = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
        top_font = Font(name="宋体", size=24, bold=True, color="000000")

        # 对齐方式
        align_global = Alignment(horizontal="center", vertical="center", wrapText=True)

        # 边框
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 需重点关注标颜色的列标题
        check_names = ["最小集", "要求类", "建议类"]
        check_names1 = [
            "不安全函数（未处理）",
            "冗余代码（未处理）",
            "超大源代码文件（未处理）",
            "超大函数（未处理）",
            "超大圈复杂度（未处理）",
            "超大目录（未处理）",
            "超大深度（未处理）",
            "抑制告警数",
        ]

        # 加载工作簿
        workb = load_workbook(cls.Libing_xlsx)
        # 选择工作表
        ws = workb["Mysheet"]
        # 打印一下当前表格的行数和列数
        log.info("Mysheet表的列数：{}".format(ws.max_column))
        log.info("Mysheet表的行数：{}".format(ws.max_row))

        # 为所有单元格 设置格式
        for row_num in range(1, ws.max_row + 1):  # 行循环
            # 行高设置
            ws.row_dimensions[row_num].height = "15.5"
            for col_num in range(1, ws.max_column + 1):  # 列循环
                # 组成单元格cell_name,例如：A5
                column_num_letter = str(get_column_letter(col_num))  # 将列数字转换为字符串
                row_num = str(row_num)
                cell_name = column_num_letter + row_num
                # 设置每个单元格的字体：微软雅黑，10号，颜色黑色
                ws[cell_name].font = Font(name="微软雅黑", size=10, color=colors.BLACK)  # 将表格字体设置为微软雅黑10号
                ws[cell_name].alignment = align_global  # 将表格字体设置为微软雅黑10号
                ws[cell_name].border = thin_border

        for col_num in range(1, ws.max_column + 1):  # 列循环
            for row_num in range(2, ws.max_row + 1):  # 行循环
                # 组成单元格cell_name,例如：A5
                column_num_letter = str(get_column_letter(col_num))
                cell_name = column_num_letter + str(row_num)

                # 模块名称列 单元格字体样式：微软雅黑，10号，单下划线，蓝色
                if ws.cell(1, col_num).value == "模块名称":
                    # ws[cell_name].font = Font(name="微软雅黑", size=10, underline="single", color="0000FF")
                    ws[cell_name].font = Font(name="微软雅黑", size=10, color="0000FF")

                # 代码覆盖率列 满足条件的单元格字体样式: 微软雅黑，10号，红色...
                if ws.cell(1, col_num).value == "代码覆盖率":
                    if ws.cell(row_num, col_num).value != "-":
                        if ws.cell(row_num, col_num).value == "error" or float(ws.cell(row_num, col_num).value) < 60:
                            ws[cell_name].font = warning_value

                # 需标颜色的列check_names
                for check_name in check_names:
                    if ws.cell(1, col_num).value == check_name:
                        if ws.cell(row_num, col_num).value != "-":
                            if (
                                ws.cell(row_num, col_num).value == "error"
                                or float(ws.cell(row_num, col_num).value) > 120272
                            ):
                                ws[cell_name].font = warning_value

                # 需标颜色的列check_names1
                for check_name in check_names1:
                    if ws.cell(1, col_num).value == check_name:
                        if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 0:
                            ws[cell_name].font = warning_value

                # 平均函数代码行列 单元格格式
                if ws.cell(1, col_num).value == "平均函数代码行":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 30:
                        ws[cell_name].font = warning_value

                # 平均文件代码行 单元格格式
                if ws.cell(1, col_num).value == "平均文件代码行":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 300:
                        ws[cell_name].font = warning_value

                # 总文件重复率
                if ws.cell(1, col_num).value == "总文件重复率":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 0.5:
                        ws[cell_name].font = warning_value

                if ws.cell(1, col_num).value == "总代码重复率":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 5:
                        ws[cell_name].font = warning_value

                if ws.cell(1, col_num).value == "平均圈复杂度":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 5:
                        ws[cell_name].font = warning_value

                if ws.cell(1, col_num).value == "最大圈复杂度":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 15:
                        ws[cell_name].font = warning_value

                if ws.cell(1, col_num).value == "架构指数":
                    if ws.cell(row_num, col_num).value != "-" and float(ws.cell(row_num, col_num).value) > 90:
                        ws[cell_name].font = warning_value

                if ws.cell(1, col_num).value == "出包":
                    if ws.cell(row_num, col_num).value == "error":
                        ws[cell_name].font = warning_value

        # 插入一行做题目
        ws.insert_rows(1, 1)
        # 先设置刚插入第一行的所有单元格的边框
        for i in range(1, ws.max_column):
            ws.cell(1, i).border = thin_border
        # 合并单元格
        ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")
        ws.cell(1, 1).value = f"{cls.product_version} {cls.y_m_d}安全扫描报告"
        # 合并后参考左上角的单元格，设置字体和对齐方式
        ws["A1"].font = top_font
        ws["A1"].alignment = align_global

        # 新增1行，超链接公式
        new_row_num = ws.max_row + 1
        new_col_letter = get_column_letter(ws.max_column)
        ws.append([])
        ws.merge_cells(f"A{new_row_num}:{new_col_letter}{new_row_num}")
        ws.cell(new_row_num, 1).value = (
            '=HYPERLINK("https://www.tenxun.com/cloud/","tip1:codelcheck列数据详情(搜索工程名即可查看数据)：https://www.tenxun.com/cloud/")'
        )

        # 新增1行，超链接公式
        new_row_num = ws.max_row + 1
        new_col_letter = get_column_letter(ws.max_column)
        ws.append([])
        ws.merge_cells(f"A{new_row_num}:{new_col_letter}{new_row_num}")
        ws.cell(new_row_num, 1).value = (
            '=HYPERLINK("https://www.tenxunyun.com/cloud/,"tip2: codelcheck后面的列为libing的数据，详情见：https://www.tenxunyun.com/cloud/")'
        )

        # 行高,列宽最后调整
        ws.row_dimensions[1].height = "60"
        ws.row_dimensions[2].height = "55"
        ws.column_dimensions["B"].width = "26.63"  # B列宽度调整

        # 保存工作簿
        workb.save(cls.Libing_xlsx)

    @classmethod
    def send_email(cls):
        """使用outlook发送邮件"""
        outlook = win32.Dispatch("outlook.Application")
        mail = outlook.CreateItem(0)

        # 发件人,默认使用outlook中的默认邮箱。如有需要使用outlook邮箱中的其他账号，如下设置
        # mail.SentOnBehalfOfName = "daiminghong@h-partners.com"

        mail.To = "xxx@qq.com"  # 收件人
        mail.CC = "xxx@qq.com"  # 抄送人
        # mail.Bcc = "123456789@qq.com" # 密抄送人

        mail.Subject = f"{cls.product_version} {cls.y_m_d}安全扫描报告"  # 邮件主题
        # mail.Importance = 2  # 设置重要性为高

        # mail.Body = "这是一封测试邮件"  # 邮件正文
        html_file = cls.transform_xlsx_to_html()
        with os.fdopen(os.open(html_file, os.O_RDONLY, stat.S_IWUSR | stat.S_IRUSR), "r", encoding="utf-8") as fs:
            Mysheet_html = fs.read()

        mail.HTMLBody = re.sub(r"<h2.*/h2>", "", Mysheet_html)  # html邮件正文

        mail.Attachments.Add(f"d:\\desktop\\{cls.Libing_xlsx}")  # 添加附件

        # 显示邮件发送界面
        mail.Display()
        # 发送
        mail.Send()

    @classmethod
    def transform_xlsx_to_html(cls):
        """把excel工作表转为html文件"""
        html_file = "Mysheet.html"
        # 创建Workbook对象
        wb = Workbook()
        # 载入Excel文件
        wb.LoadFromFile(cls.Libing_xlsx)
        # 获取工作表
        sheet = wb.Worksheets.get_Item(0)
        # 将工作表转为html文件
        sheet.SaveToHtml(html_file)
        wb.Dispose()
        return html_file


if __name__ == "__main__":
    # Daily_Email.process_libing_xlsx()
    # Daily_Email.beautiful_libing_xlsx()
    # Daily_Email.transform_xlsx_to_html()
    # Daily_Email.send_email()
    pass
